# Servicio de generacion de reportes: crea archivos PDF, Excel y CSV con el listado de facturas
# Cada formato usa una libreria distinta: fpdf2 para PDF, openpyxl para Excel, csv nativo para CSV

import csv
from datetime import datetime
from fpdf import FPDF
import openpyxl
from sqlalchemy.orm import Session
from app.models.factura import Factura

def _obtener_facturas(db: Session):
    # Consulta todas las facturas ordenadas de la mas reciente a la mas antigua
    return db.query(Factura).order_by(Factura.creado_en.desc()).all()

def generar_pdf(db: Session, ruta: str):
    # Genera un reporte PDF con tabla de facturas usando la libreria fpdf2
    facturas = _obtener_facturas(db)
    pdf = FPDF()
    pdf.add_page()

    # Titulo centrado del reporte
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "SmartInvoice - Reporte de Facturas", ln=True, align="C")

    # Subtitulo con fecha y hora de generacion en UTC
    pdf.set_font("Helvetica", size=9)
    pdf.cell(0, 6, f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", ln=True, align="C")
    pdf.ln(5)

    # Definicion de columnas: nombre y ancho en mm
    headers = ["ID", "N Factura", "Fecha", "Proveedor ID", "Subtotal", "Impuestos", "Total", "Estado"]
    col_widths = [10, 35, 22, 25, 22, 22, 22, 22]

    # Encabezado de tabla con fondo azul y texto blanco
    pdf.set_fill_color(30, 90, 160)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for h, w in zip(headers, col_widths):
        pdf.cell(w, 8, h, border=1, fill=True)
    pdf.ln()

    # Filas de datos con colores alternados para mejor legibilidad
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", size=8)
    for i, f in enumerate(facturas):
        fill = i % 2 == 0   # Filas pares con fondo celeste claro, impares con blanco
        pdf.set_fill_color(240, 245, 255) if fill else pdf.set_fill_color(255, 255, 255)
        row = [
            str(f.id),
            f.numero_factura or "-",
            str(f.fecha_factura) if f.fecha_factura else "-",
            str(f.proveedor_id) if f.proveedor_id else "-",
            f"{f.subtotal:.2f}" if f.subtotal else "-",
            f"{f.impuestos:.2f}" if f.impuestos else "-",
            f"{f.total:.2f}" if f.total else "-",
            f.estado,
        ]
        for val, w in zip(row, col_widths):
            pdf.cell(w, 7, val, border=1, fill=fill)
        pdf.ln()

    # Guarda el archivo PDF en la ruta especificada
    pdf.output(ruta)

def generar_excel(db: Session, ruta: str):
    # Genera un archivo Excel con una hoja de trabajo usando openpyxl
    facturas = _obtener_facturas(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # Primera fila con nombres de columnas
    headers = ["ID", "N Factura", "Fecha", "Proveedor ID", "Subtotal", "Impuestos", "Total", "Estado", "Creado en"]
    ws.append(headers)

    # Agrega una fila por cada factura con sus valores correspondientes
    for f in facturas:
        ws.append([
            f.id,
            f.numero_factura,
            str(f.fecha_factura) if f.fecha_factura else None,
            f.proveedor_id,
            float(f.subtotal) if f.subtotal else None,
            float(f.impuestos) if f.impuestos else None,
            float(f.total) if f.total else None,
            f.estado,
            str(f.creado_en),
        ])

    wb.save(ruta)

def generar_csv(db: Session, ruta: str):
    # Genera un archivo CSV con valores separados por comas, codificado en UTF-8
    facturas = _obtener_facturas(db)
    with open(ruta, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        # Encabezados en la primera fila
        writer.writerow(["ID", "N Factura", "Fecha", "Proveedor ID", "Subtotal", "Impuestos", "Total", "Estado", "Creado en"])
        # Una fila por factura
        for f in facturas:
            writer.writerow([
                f.id, f.numero_factura, f.fecha_factura, f.proveedor_id,
                f.subtotal, f.impuestos, f.total, f.estado, f.creado_en,
            ])
